import argparse
import sys
from bs4 import BeautifulSoup
import requests
from queue import Queue
from urllib.parse import urljoin, urlparse
from openpyxl import Workbook
import subprocess
import platform

# Define the function to parse command line arguments
def parse_arguments():
    parser = argparse.ArgumentParser(description="Crawl websites and save internal links ending with .html to an Excel file with multiple sheets.",
                                     formatter_class=argparse.RawDescriptionHelpFormatter,
                                     epilog="Usage examples:\n"
                                            "1. To specify a sites file: python script.py --sites_file sites.txt\n"
                                            "2. To provide base URLs directly: python script.py https://example.com https://anotherexample.com")

    group = parser.add_argument_group("Input Options")
    group.add_argument("--sites_file", help="Path to a text file containing a list of websites")
    group.add_argument("base_urls", nargs="*", help="Base URLs of the websites to crawl", default=[])

    # If no arguments are provided, display the help message
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(1)

    args = parser.parse_args()

    return args

# Parse command line arguments
args = parse_arguments()

# Create an Excel file
workbook = Workbook()

# Determine the list of websites to crawl
sites_from_file = []
if args.sites_file:
    with open(args.sites_file, "r") as file:
         sites_from_file = [line.strip() for line in file.readlines()]

if args.base_urls:
    sites_to_crawl = args.base_urls + sites_from_file
else:
    sites_to_crawl = sites_from_file

# Define a function to process and collect data from web pages
def process_site(base_url, worksheet):
    url_queue = Queue()
    url_queue.put(base_url)
    visited_urls = set()

    while not url_queue.empty():
        url = url_queue.get()
        if url not in visited_urls:
            try:
                response = requests.get(url)
                soup = BeautifulSoup(response.text, "html.parser")
                for link in soup.find_all("a"):
                    href = link.get("href")
                    text = link.get_text()
                    if href is not None and not href.startswith("#"):
                        absolute_url = urljoin(url, href)
                        if base_url in absolute_url and absolute_url.endswith(".html"):
                            if url != absolute_url:
                                if not url.endswith(".html"):
                                    worksheet.append([url, absolute_url, text])
                        if urlparse(absolute_url).netloc == urlparse(base_url).netloc:
                            if absolute_url not in visited_urls:
                                url_queue.put(absolute_url)
                visited_urls.add(url)
            except Exception as e:
                print(f"Processing error {url}: {str(e)}")

# Create a dictionary to store worksheets for each site
site_worksheets = {}

# Process each site and collect data
for base_url in sites_to_crawl:
    # Create a new worksheet for the site
    site_title = base_url.split("//")[-1].replace("/", "_")
    if len(site_title) > 30:
        site_title = site_title[:30]
    worksheet = workbook.create_sheet(title=site_title)
    site_worksheets[base_url] = worksheet

    # Add column headers
    worksheet.append(["Source URL", "Target URL", "Link Text"])

    # Call the function to process and collect data from the web pages
    process_site(base_url, worksheet)

for sheet in workbook.sheetnames:
    if sheet == "Sheet":
        workbook.remove(workbook[sheet])

# Save the Excel file
workbook.save("Results.xlsx")

print("Exported to Results.xlsx")

# Check if the operating system is Windows (you can adapt this for other systems)
if platform.system() == "Windows":
    # Open the file with Microsoft Excel
    command = f'start excel "Results.xlsx"'
    
    # Run the command
    subprocess.Popen(command, shell=True)
